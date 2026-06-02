"""
src/pipeline/nodes/report_generator.py

Report Generator — called ONCE by the runner loop after all variants are processed.
NOT a LangGraph node — this is a standalone function that receives the full list of
completed VariantState dicts and produces output files.

Usage (in your runner loop):
    from src.pipeline.nodes.report_generator import generate_reports

    completed_states = []
    for variant in variants:
        state = build_initial_state(...)
        result = VARIANT_GRAPH.invoke(state)
        completed_states.append(result)

    paths = generate_reports(
        states=completed_states,
        session_id=session_id,
        output_dir=Path(work_dir) / "reports",
        formats=config.output_formats,          # ["xlsx", "tsv", "html"] or subset
        report_config=REPORT_CONFIG,            # from config.py
    )
    # paths = {"xlsx": Path(...), "tsv": Path(...), "html": Path(...)}

Design decisions vs the guide:
  - Runner-loop accumulation: generator receives ALL states at once, not one at a time.
    Avoids shared-file race conditions; enables summary cards and cross-variant sorting.
  - Jinja2 HTML template (not df.to_html): supports letterhead, conditional blocks,
    expandable rows, and print CSS.
  - TSV is always written (zero dependencies), xlsx and html are opt-in via `formats`.
  - All paths are returned so the FastAPI layer can serve them directly.
"""

import logging
from dataclasses import dataclass, field
from pathlib import Path
from datetime import datetime
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from jinja2 import Environment, FileSystemLoader, select_autoescape

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# ACMG classification ordering and colours
# ---------------------------------------------------------------------------

CLASSIFICATION_ORDER = {
    "Pathogenic":        1,
    "Likely_Pathogenic": 2,
    "VUS":               3,
    "Likely_Benign":     4,
    "Benign":            5,
}

# xlsx fill colours (ARGB hex, no alpha prefix needed for openpyxl)
XLSX_FILL_COLORS = {
    "Pathogenic":        "FFCDD2",   # soft red
    "Likely_Pathogenic": "FFE0B2",   # soft orange
    "VUS":               "FFF9C4",   # soft yellow
    "Likely_Benign":     "DCEDC8",   # soft lime
    "Benign":            "C8E6C9",   # soft green
}

XLSX_FONT_COLORS = {
    "Pathogenic":        "B71C1C",
    "Likely_Pathogenic": "E65100",
    "VUS":               "F57F17",
    "Likely_Benign":     "558B2F",
    "Benign":            "1B5E20",
}


# ---------------------------------------------------------------------------
# Report configuration — pass in from config.py or override per-session
# ---------------------------------------------------------------------------

@dataclass
class ReportConfig:
    """
    Controls branding and optional content in HTML reports.
    Populate from your PipelineConfig / environment variables.
    """
    lab_name:         str = "Genomics Laboratory"
    lab_subtitle:     str = ""                            # e.g. department name
    lab_contact:      str = ""                            # e.g. email / phone
    logo_path:        Optional[str] = None                # absolute path to PNG/SVG
    pipeline_version: str = "2.0"
    genome_build:     str = "GRCh38"
    disclaimer:       str = (
        "This report is intended for clinical research use only. "
        "Variant classifications should be interpreted by a qualified clinical "
        "geneticist in the context of the patient's full clinical presentation."
    )
    # Classification tiers to include in report (set to subset to filter output)
    include_classes:  list = field(default_factory=lambda: [
        "Pathogenic", "Likely_Pathogenic", "VUS", "Likely_Benign", "Benign"
    ])


# Default instance — import and override as needed
DEFAULT_REPORT_CONFIG = ReportConfig()


# ---------------------------------------------------------------------------
# Column definitions — single source of truth for both xlsx and tsv
# ---------------------------------------------------------------------------

# Primary columns shown in the main report table
PRIMARY_COLUMNS = [
    ("Rank",                "rank"),
    ("Variant",             "variant_id"),
    ("Gene",                "gene"),
    ("HGVSc",               "hgvsc"),
    ("HGVSp",               "hgvsp"),
    ("Consequence",         "consequence"),
    ("Classification",      "final_classification"),
    ("Criteria",            "criteria_applied"),
    ("Confidence",          "confidence"),
    ("Phenotype Score",     "phenotype_score"),
    ("gnomAD AF",           "max_gnomad_af"),
    ("ClinVar",             "clinvar_clnsig"),
    ("ClinVar ★",          "clinvar_stars"),
]

# Detail columns shown in expandable rows (HTML) / extra sheet (xlsx)
DETAIL_COLUMNS = [
    ("Evidence Summary",        "evidence_summary"),
    ("REVEL",                   "revel_score"),
    ("SpliceAI",                "max_spliceai"),
    ("CADD PHRED",              "cadd_phred"),
    ("Phase Status",            "phase_status"),
    ("Zygosity Filter",         "zygosity_filter_status"),
    ("Matched Disease",         "matched_orphanet_disease"),
    ("Orphanet ID",             "orphanet_id"),
    ("HPO Matched Genes",       "hpo_matched_genes_str"),
    ("Debate Notes",            "debate_notes"),
    ("Unevaluated Criteria",    "unevaluated_str"),
    ("Recommended Followup",    "recommended_followup"),
    ("Reclassification If",     "reclassification_conditions"),
    ("Zygosity Filter",         "zygosity_filter_status"),
    ("Phase Confidence",        "phase_confidence"),
    ("Warnings",                "warnings_str"),
    ("Citations",               "citations_str"),
]


# ---------------------------------------------------------------------------
# State → flat row conversion
# ---------------------------------------------------------------------------

def _state_to_row(state: dict, rank: int) -> dict:
    """
    Convert a completed VariantState dict to a flat report row.
    All list/dict fields are serialised to strings here.
    """
    unevaluated = state.get("unevaluated_criteria_report") or []
    warnings    = state.get("warnings") or []
    citations   = state.get("all_citations") or []
    hpo_genes   = state.get("hpo_matched_genes") or []

    # Criteria: use final_criteria_applied (post-debate definitive list)
    criteria = state.get("final_criteria_applied") or []

    phenotype_score = state.get("phenotype_score")
    gnomad_af       = state.get("max_gnomad_af", 0.0)

    return {
        # --- primary ---
        "rank":                 rank,
        "variant_id":           state.get("variant_id", ""),
        "gene":                 state.get("gene", ""),
        "hgvsc":                state.get("hgvsc") or "",
        "hgvsp":                state.get("hgvsp") or "",
        "consequence":          state.get("consequence", ""),
        "final_classification": state.get("final_classification") or "VUS",
        "criteria_applied":     ", ".join(criteria) if criteria else "—",
        "confidence":           state.get("confidence") or "LOW",
        "phenotype_score":      f"{phenotype_score:.3f}" if phenotype_score is not None else "—",
        "max_gnomad_af":        f"{gnomad_af:.6f}" if gnomad_af else "0.000000",
        "clinvar_clnsig":       state.get("clinvar_clnsig") or "—",
        "clinvar_stars":        state.get("clinvar_stars", 0),

        # --- detail ---
        "evidence_summary":           state.get("evidence_summary") or "",
        "revel_score":                state.get("revel_score"),
        "max_spliceai":               state.get("max_spliceai"),
        "cadd_phred":                 state.get("cadd_phred"),
        "phase_status":               state.get("phase_status") or "",
        "phase_confidence":           state.get("phase_confidence") or "",
        "zygosity_filter_status":     state.get("zygosity_filter_status") or "",
        "matched_orphanet_disease":   state.get("matched_orphanet_disease") or "",
        "orphanet_id":                state.get("orphanet_id") or "",
        "hpo_matched_genes_str":      ", ".join(hpo_genes) if hpo_genes else "—",
        "debate_notes":               state.get("debate_notes") or "",
        "unevaluated_str":            ", ".join(unevaluated) if unevaluated else "None",
        "recommended_followup":       state.get("recommended_followup") or "",
        "reclassification_conditions":state.get("reclassification_conditions") or "",
        "warnings_str":               "; ".join(warnings) if warnings else "",
        "citations_str":              "; ".join(citations) if citations else "",

        # raw state reference (for HTML template logic)
        "_has_unevaluated": bool(unevaluated),
        "_has_warnings":    bool(warnings),
        "_classification":  state.get("final_classification") or "VUS",
        "_session_id":      state.get("session_id", ""),
    }


def _build_rows(states: list, include_classes: list) -> list:
    """
    Convert all states to rows, sort by classification tier then phenotype score,
    assign ranks, and filter to requested classification tiers.
    """
    rows = []
    for state in states:
        cls = state.get("final_classification") or "VUS"
        if cls not in include_classes:
            continue
        rows.append(state)

    # Sort: classification tier (P first) then phenotype_score descending
    def sort_key(s):
        cls   = s.get("final_classification") or "VUS"
        score = s.get("phenotype_score") or 0.0
        return (CLASSIFICATION_ORDER.get(cls, 99), -score)

    rows.sort(key=sort_key)

    return [_state_to_row(s, i + 1) for i, s in enumerate(rows)]


def _classification_counts(rows: list) -> dict:
    """Summary counts for HTML header cards."""
    counts = {k: 0 for k in CLASSIFICATION_ORDER}
    for row in rows:
        cls = row.get("_classification", "VUS")
        if cls in counts:
            counts[cls] += 1
    counts["total"] = len(rows)
    return counts


# ---------------------------------------------------------------------------
# TSV writer
# ---------------------------------------------------------------------------

def _write_tsv(rows: list, path: Path):
    all_cols = [c[1] for c in PRIMARY_COLUMNS] + [c[1] for c in DETAIL_COLUMNS]
    # remove internal keys
    all_cols = [c for c in all_cols if not c.startswith("_")]
    df = pd.DataFrame(rows)[all_cols]
    df.to_csv(str(path), sep="\t", index=False)
    logger.info(f"TSV written: {path}")


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------

def _write_xlsx(rows: list, path: Path, session_id: str, report_config: ReportConfig):
    """
    Two-sheet Excel workbook:
      Sheet 1 — Summary (primary columns, colour-coded classification)
      Sheet 2 — Full Detail (all columns, one row per variant)
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "ACMG Summary"

    # Title rows
    ws1.merge_cells("A1:M1")
    title_cell = ws1["A1"]
    title_cell.value = f"{report_config.lab_name} — ACMG Variant Classification Report"
    title_cell.font = Font(bold=True, size=13, color="1A237E")
    title_cell.alignment = Alignment(horizontal="center")

    ws1.merge_cells("A2:M2")
    meta_cell = ws1["A2"]
    meta_cell.value = (
        f"Session: {session_id}  |  "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"Build: {report_config.genome_build}  |  "
        f"Pipeline v{report_config.pipeline_version}  |  "
        f"Variants: {len(rows)}"
    )
    meta_cell.font = Font(italic=True, size=9, color="555555")
    meta_cell.alignment = Alignment(horizontal="center")

    # Header row
    header_row = 4
    header_labels = [c[0] for c in PRIMARY_COLUMNS]
    thin = Side(border_style="thin", color="CCCCCC")
    header_border = Border(bottom=thin)

    for col_idx, label in enumerate(header_labels, start=1):
        cell = ws1.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A237E")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = header_border

    # Data rows
    primary_keys = [c[1] for c in PRIMARY_COLUMNS]
    for row_idx, row in enumerate(rows, start=header_row + 1):
        cls = row.get("_classification", "VUS")
        fill_color = XLSX_FILL_COLORS.get(cls, "FFFFFF")
        font_color = XLSX_FONT_COLORS.get(cls, "000000")

        for col_idx, key in enumerate(primary_keys, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(
                bottom=Side(border_style="thin", color="E0E0E0"),
                right=Side(border_style="thin",  color="E0E0E0"),
            )
            # Colour only the classification column
            if key == "final_classification":
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(bold=True, color=font_color)

        # Unevaluated criteria warning marker in Criteria column
        if row.get("_has_unevaluated"):
            crit_col = primary_keys.index("criteria_applied") + 1
            existing = ws1.cell(row=row_idx, column=crit_col).value or ""
            ws1.cell(row=row_idx, column=crit_col).value = existing + "  ⚠ PP4/BP5 not evaluated"
            ws1.cell(row=row_idx, column=crit_col).font = Font(color="E65100")

    # Column widths for Sheet 1
    col_widths = [6, 22, 10, 24, 20, 20, 18, 30, 10, 14, 12, 16, 8]
    for i, w in enumerate(col_widths, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws1.row_dimensions[1].height = 20
    ws1.freeze_panes = ws1.cell(row=header_row + 1, column=1)

    # ── Sheet 2: Full Detail ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Full Detail")
    all_col_defs = PRIMARY_COLUMNS + DETAIL_COLUMNS
    all_col_defs = [c for c in all_col_defs if not c[1].startswith("_")]
    all_keys = [c[1] for c in all_col_defs]

    for col_idx, (label, _) in enumerate(all_col_defs, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="37474F")
        cell.alignment = Alignment(wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(all_keys, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=Side(border_style="thin", color="EEEEEE"))

    for i in range(1, len(all_col_defs) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 22
    ws2.freeze_panes = ws2["A2"]

    wb.save(str(path))
    logger.info(f"XLSX written: {path}")


# ---------------------------------------------------------------------------
# HTML writer — Jinja2 template
# ---------------------------------------------------------------------------

def _logo_to_base64(logo_path: Optional[str]) -> Optional[str]:
    """Inline logo as base64 data URI so HTML is self-contained."""
    if not logo_path:
        return None
    try:
        import base64
        p = Path(logo_path)
        ext = p.suffix.lower().lstrip(".")
        mime = {"png": "image/png", "svg": "image/svg+xml", "jpg": "image/jpeg"}.get(ext, "image/png")
        data = base64.b64encode(p.read_bytes()).decode("utf-8")
        return f"data:{mime};base64,{data}"
    except Exception as e:
        logger.warning(f"Could not inline logo {logo_path}: {e}")
        return None


def _write_html(rows: list, path: Path, session_id: str, report_config: ReportConfig):
    """Render the Jinja2 HTML template."""
    template_dir = Path(__file__).parent.parent.parent / "report_templates"
    env = Environment(
        loader=FileSystemLoader(str(template_dir)),
        autoescape=select_autoescape(["html"]),
    )
    template = env.get_template("acmg_report.html.j2")

    counts = _classification_counts(rows)
    logo_uri = _logo_to_base64(report_config.logo_path)

    html = template.render(
        session_id       = session_id,
        generated_at     = datetime.now().strftime("%Y-%m-%d %H:%M"),
        lab_name         = report_config.lab_name,
        lab_subtitle     = report_config.lab_subtitle,
        lab_contact      = report_config.lab_contact,
        logo_uri         = logo_uri,
        pipeline_version = report_config.pipeline_version,
        genome_build     = report_config.genome_build,
        disclaimer       = report_config.disclaimer,
        counts           = counts,
        rows             = rows,
        primary_columns  = PRIMARY_COLUMNS,
        detail_columns   = DETAIL_COLUMNS,
    )

    path.write_text(html, encoding="utf-8")
    logger.info(f"HTML written: {path}")


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def generate_reports(
    states:        list,
    session_id:    str,
    output_dir:    Path,
    formats:       list = None,
    report_config: ReportConfig = None,
) -> dict:
    """
    Generate output reports from a list of completed VariantState dicts.

    Args:
        states:        All completed VariantState dicts from the runner loop.
        session_id:    Session identifier (used in filenames and report header).
        output_dir:    Directory to write output files into (created if absent).
        formats:       List of formats to produce. Subset of ["xlsx", "tsv", "html"].
                       Default: all three.
        report_config: Branding/content configuration. Defaults to DEFAULT_REPORT_CONFIG.

    Returns:
        Dict mapping format name to output Path, e.g.:
        {"xlsx": Path("/output/abc123_acmg.xlsx"), "tsv": ..., "html": ...}
    """
    if formats is None:
        formats = ["xlsx", "tsv", "html"]
    if report_config is None:
        report_config = DEFAULT_REPORT_CONFIG

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if not states:
        logger.warning("generate_reports called with empty states list — no output produced.")
        return {}

    logger.info(f"Generating reports for {len(states)} variants | formats={formats}")

    rows = _build_rows(states, include_classes=report_config.include_classes)

    if not rows:
        logger.warning("No variants passed the include_classes filter.")
        return {}

    outputs = {}
    base = output_dir / f"{session_id}_acmg_report"

    if "tsv" in formats:
        tsv_path = base.with_suffix(".tsv")
        _write_tsv(rows, tsv_path)
        outputs["tsv"] = tsv_path

    if "xlsx" in formats:
        xlsx_path = base.with_suffix(".xlsx")
        _write_xlsx(rows, xlsx_path, session_id, report_config)
        outputs["xlsx"] = xlsx_path

    if "html" in formats:
        html_path = base.with_suffix(".html")
        _write_html(rows, html_path, session_id, report_config)
        outputs["html"] = html_path

    logger.info(f"Reports complete: {list(outputs.keys())}")
    return outputs
